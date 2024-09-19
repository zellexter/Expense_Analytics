# EXPENSES PROJECT
# Goal: create in depth visualization for expenses data
#       include:
#           SUMMARY
#               bar chart with monthly total exp
#               pie chart with yearly exp categories, highlight on top 3
#               line chart for each category's monthly data (grocery, dining, shopping, pet, etc.)
#               averages for each cateogory over the year
#           MONTHLY
#               exp data
#               pie chart expenditure by category
#               high low categories
#               this month compared to the previous month and compared to the average
#               Dining VS. Groceries
# Skills to practice:
#       xlsx file read in
#       data extraction on xlsx files
#       pandas df manipulation
#       creating and importing modules
#       visualization with charts (by calling external functions created previously)
#           line chart for monthly totals, by month category totals
#           pie chart for monthly category expenditure
# Data Format:
#       WHERE: EXPENSES_CHROMA.xlsx is kept within Expenses_Data Folder in CodingWDad\Expense Analytics
#       FILE AND DATA: EXPENSES_CHROMA.xlsx contains worksheets with each month's data
#       DATA DESCRIPTION: Within each worksheet, data is stored starting from row20 (header inclusive)
#                         across columns A:C where:
#                                               A: Expenditure (name)
#                                               B: Category (data validation rules)
#                                               C: Amount (in dollars)
#       NOTE*** first work on expense based data in described columns. don't worry about subscriptions/monthly expenses yet.

# BREAKDOWN:

# DONE READ IN FILE
# DONE IMPORT DATA FROM WS TO DF
# for ws in wb:
#   store data into df using min/max row/col
# DONE CLEAN DATA USING PANDAS -> DATA_CLEANER_MODULE.PY
# create a module to clean data using pandas (this would be a seperate .py file)
# TODO FORMAT DATA FOR CHARTS -> DATA_FORMATTER_MODULE.PY
# find what data format is needed for each type of chart
# create functions for each type of chart that will format df data
# TODO DRAW/SAVE CHARTS -> DATA_VISUALIZER_MODULE.PY
# DONE EXPORT NEW FILE WITH VISUALIZATIONS

# NOTE*** LATER IMPROVEMENTS
# currently the plan is to store data from file into pandas df (to practice pandas).
#   future improvement may be to read and extract data directly from file into clean/format functions
# BANK STATEMENTS---
#   pass in pdfs of bank statements and read expense data from there
#   need to learn:
#       converting pdf into txt files or something readable
#       finding right data within the txt files and assigning correct dollar amounts to the expense name
#       categorize expenses by name and send unknown expenses to manual input box, and then learn from there
#   workbook:
#       ws1: yearly overview
#       ws2+: monthly expense data with visualizations on same ws

# MODULES
import pandas as pd
from pandas import DataFrame
import os
from openpyxl import Workbook, load_workbook
from Data_Cleaner_Module import convert_float, convert_str, capitalize
from Data_Formatter_Module import *
from Data_Visualizer_Module import *


class Expense_Visualization():

    def __init__(self, import_file, export_file) -> None:
        '''Initializes import and export file, sets original workbook to None'''
        self.import_file = import_file
        self.export_file = export_file
        self.wb = None

    def main(self):
        '''Main function:
            Reads in file
            Imports data from worksheet to dataframes
            Cleans data in dataframe using DataCleaner Module
            Draw and save chart in new file
            Export new file with visualizations'''

        dict_df = self.file_import(data_type='expenses')  # {worksheet:df}
        self.wb = self.create_wb()
        dict_df_totals = {worksheet: dataframe.groupby('Category')['Amount'].sum().reset_index()
                          for worksheet, dataframe in dict_df.items()}  # {worksheet:df}

        # CLEANS DATA
        for df in dict_df.values():  # each worksheet's dataframe
            df['Amount'] = df['Amount'].apply(convert_float)
            df['Expenditure'] = df['Expenditure'].apply(convert_str)
            df['Expenditure'] = df['Expenditure'].apply(capitalize)

        # CREATE WS AND APPEND DATA
        for worksheet, dataframe in dict_df.items():
            # creates new worksheet
            ws = self.create_ws(worksheet)
            # appends column headers to first available cells
            ws.append(list(dataframe.columns))
            # appends dataframe to ws
            self.append_rows(ws, dataframe)

        # CREATE PIECHART FOR MONTHLY CATEGORY DATA
        self.create_monthly_piechart(dict_df_totals)
            

        self.file_export()

    def create_monthly_piechart(self, dict_df_totals):
        '''Creates pie chart for each month's total expenses seperated by category'''
        for worksheet_title, dataframe in dict_df_totals.items():
            header = ['Expense Category', 'Amount']
            worksheet = self.wb[worksheet_title]

            # appending reference data
            start_row = 1 # start from first row
            start_col = 4 # start from fourth column D

            # title
            title = f'{worksheet_title} Expenses by Category'

            draw_piechart(title, worksheet, header, dataframe, start_row, start_col, add_to='F2')

    # TODO: def create_summary_piechart(self, 

    def append_rows(self, ws, df: DataFrame):
        '''Iterates through rows in df and appends to ws'''
        for row in df.itertuples(index=False, name=None):
            ws.append(row)

    def file_import(self, data_type: str = 'all') -> dict[str, DataFrame]:
        '''
        Imports all worksheets to dataframes.
        Dataframes stored in dictionary {worksheet:dataframe}.
        data_type parameter defaulted to import all data in worksheet(s),
        however can import only 'fixed' and only 'expense' data as well.

        Returns 
        '''
        # TODO do not import template ws
        # TODO fix fixed expenses, set amount of rows
        # DONE simplify!!! no need for all the if statements, can store type:parameters in a dictionary
        #   initialize dict for import_parameters
        #   initialize variable parameters to get() parameters by data_type
        #   return read_excel function with parameters unpacked

        import_parameters = {
            'all': {},
            'fixed': {
                'header': None,
                'names': ['Fixed Expense', 'Type', 'Amount'],
                'usecols': 'A:C',
                'skiprows': 5,
                'skipfooter': 18
            },
            'expenses': {
                'header': None,
                'names': ['Expenditure', 'Category', 'Amount'],
                'usecols': 'A:C',
                'skiprows': 20
            }
        }

        parameters = import_parameters.get(data_type, {})

        return pd.read_excel(io=self.import_file, sheet_name=None, **parameters)

    def file_export(self) -> None:
        '''Saves into export file as xlsx'''
        self.wb.save(self.export_file)

    def create_wb(self) -> Workbook:
        '''Creates a new workbook'''
        wb = Workbook()

        # retitle default sheet to Summary
        default_ws = wb['Sheet']
        default_ws.title = 'Summary'

        return wb

    def create_ws(self, ws_name: str):
        '''Creates a new worksheet where ws_name is the name of the worksheet'''
        return self.wb.create_sheet(ws_name)


# if Expense_Visuatlization is run on original py file, use specified filepath and save_file
if __name__ == '__main__':
    filepath = os.path.join('C:\\', 'Users', 'miche', 'OneDrive', 'Documents',
                            'CodingwDad', 'Expense_Analytics', 'Expense_Data', 'EXPENSES_CHROMA (3).xlsx')
    save_file = os.path.join('C:\\', 'Users', 'miche', 'OneDrive', 'Documents',
                             'CodingwDad', 'Expense_Analytics', 'Expense_Summary.xlsx')
    Expense_Visualization(import_file=filepath, export_file=save_file).main()


#  def foobar(self, foo:str = 'FOO') -> str:
#         return foo + '_bar'

#     def foobar_old(self, foo='FOO'):
#         return foo + '_bar'

#     def foobar_general(self, foo, bar):
#         return foo + bar

#     def more_stuff(self, x:str, y:str) -> bool:
#         return x == y


#     # forbar()

#     # char * foobar(char* foo, char* bar):

#     more_stuff(foobar('xyz'), foobar(1), 'aaa')\

# coverage term line python3 -m pytest --cov-report term --cov-report xml:coverage.xml --cov=.

