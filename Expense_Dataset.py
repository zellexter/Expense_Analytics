'''
This document extracts and stores expenses data in dataset.xlsx
The exported file organizes information into the following columns:
    - Expenditure
    - Category
    - Amount
    - Month/Year
'''

import pandas as pd
from pandas import DataFrame
from Data_Cleaner_Module import convert_float, convert_str, capitalize
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import numbers
import datetime

import_fp = os.path.join('C:\\', 'Users', 'miche', 'OneDrive', 'Documents',
                            'CodingwDad', 'Expense_Analytics', 'Expense_Data', 'EXPENSES_CHROMA.xlsx')
export_fp = os.path.join('C:\\', 'Users', 'miche', 'OneDrive', 'Documents',
                            'CodingwDad', 'Expense_Analytics', 'dataset.xlsx')

# IMPORT -> dict[str, DataFrame]
dict_df = pd.read_excel(
    io= import_fp,
    sheet_name=None,
    header=None,
    names=['Expenditure', 'Category', 'Amount'],
    usecols='A:C',
    skiprows=20,
    )
del dict_df['TEMPLATE']

# CREATE WB
wb = Workbook()
default_ws = wb['Sheet']
default_ws.title = 'Expenses_Dataset'

# DATAFRAME
for ws,df in dict_df.items():
    df['Month/Year'] = [ws]* len(df.index)

# CLEAN DF
for df in dict_df.values():
            df['Amount'] = df['Amount'].apply(convert_float)
            df['Expenditure'] = df['Expenditure'].apply(convert_str)
            df['Expenditure'] = df['Expenditure'].apply(capitalize)
            df['Month/Year'] = df['Month/Year'].apply(lambda x: datetime.datetime.strptime(x, '%b%y'))

# APPEND
headers = ['Expenditure', 'Category', 'Amount', 'Month/Year']
default_ws.append(headers)
for ws, df in dict_df.items():
    for row in df.itertuples(index=False, name=None):
        default_ws.append(row)

# CLEAN XLSX
for row in default_ws.iter_rows(min_row=2, max_row=default_ws.max_row, min_col=3,max_col=3):
    for cell in row:
          cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

# EXPORT
wb.save(export_fp)